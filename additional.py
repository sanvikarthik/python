#!/usr/bin/env python
# coding: utf-8

# In[1]:


#prog1
# Get the student's score as input
score = float(input("Enter the student's score: "))

# Determine the grade and print a message
if score >= 90:
    grade = 'A'
    message = 'Excellent!'
elif score >= 80:
    grade = 'B'
    message = 'Very Good!'
elif score >= 70:
    grade = 'C'
    message = 'Good'
elif score >= 60:
    grade = 'D'
    message = 'Needs Improvement'
else:
    grade = 'F'
    message = 'Fail'

print(f'Grade: {grade}')
print(f'Message: {message}')


# #prog2
# 

# In[2]:


#prog2
# Get two numbers and the operator as input
num1 = float(input("Enter the first number: "))
operator = input("Enter the operator (+, -, *, /): ")
num2 = float(input("Enter the second number: "))

# Perform the calculation based on the operator
if operator == '+':
    result = num1 + num2
elif operator == '-':
    result = num1 - num2
elif operator == '*':
    result = num1 * num2
elif operator == '/':
    if num2 == 0:
        result = "Division by zero is not allowed"
    else:
        result = num1 / num2
else:
    result = "Invalid operator"

print(f'Result: {result}')


# In[3]:


#progg3
# Find the largest of two numbers
def find_largest_two(a, b):
    if a > b:
        return a
    else:
        return b

# Find the largest of three numbers
def find_largest_three(a, b, c):
    largest_two = find_largest_two(a, b)
    return find_largest_two(largest_two, c)

# Get input from the user
num1 = float(input("Enter the first number: "))
num2 = float(input("Enter the second number: "))
num3 = float(input("Enter the third number: "))

# Find and print the largest number
largest = find_largest_three(num1, num2, num3)
print(f"The largest number is: {largest}")


# In[6]:


#prog4
# Function to check if a number is an Armstrong number
def is_armstrong(number):
    num_str = str(number)
    num_digits = len(num_str)
    sum_of_cubes = sum(int(digit) ** num_digits for digit in num_str)
    return sum_of_cubes == number

# Get input from the user
num = int(input("Enter a number: "))

# Check and print if it's an Armstrong numberji
if is_armstrong(num):
    print(f"{num} is an Armstrong number")
else:
    print(f"{num} is not an Armstrong number")


# In[7]:


#prog5
# Import the complex math module for handling complex roots
import cmath

# Get coefficients from the user
a = float(input("Enter the coefficient 'a': "))
b = float(input("Enter the coefficient 'b': "))
c = float(input("Enter the coefficient 'c': "))

# Calculate the discriminant
discriminant = (b ** 2) - (4 * a * c)

# Find two solutions
root1 = (-b - cmath.sqrt(discriminant)) / (2 * a)
root2 = (-b + cmath.sqrt(discriminant)) / (2 * a)

# Print the solutions
print(f"Root 1: {root1}")
print(f"Root 2: {root2}")


# In[8]:


#prog6
# Get input from the user
n = int(input("Enter a positive integer 'n': "))

# Calculate the sum of natural numbers from 1 to n
sum_of_numbers = (n * (n + 1)) // 2

# Calculate the average
average = sum_of_numbers / n

# Print the sum and average
print(f"Sum of the first {n} natural numbers: {sum_of_numbers}")
print(f"Average of the first {n} natural numbers: {average}")


# 

# In[9]:


#prog7
# Function to check if a number is prime
def is_prime(num):
    if num <= 1:
        return False
    for i in range(2, int(num**0.5) + 1):
        if num % i == 0:
            return False
    return True

# Get input from the user
num = int(input("Enter a number: "))

# Check and print if it's prime
if is_prime(num):
    print(f"{num} is a prime number")
else:
    print(f"{num} is not a prime number")


# In[10]:


#prog8
# Function to calculate the factorial of a number
def factorial(n):
    if n == 0:
        return 1
    else:
        return n * factorial(n - 1)

# Get input from the user
num = int(input("Enter a non-negative integer: "))

# Check if the input is non-negative
if num < 0:
    print("Factorial is not defined for negative numbers.")
else:
    result = factorial(num)
    print(f"The factorial of {num} is {result}")


# In[11]:


#prog9
# Get the temperature in Celsius from the user
celsius = float(input("Enter temperature in Celsius: "))

# Convert Celsius to Fahrenheit
fahrenheit = (celsius * 9/5) + 32

# Print the result
print(f"{celsius} degrees Celsius is equal to {fahrenheit} degrees Fahrenheit")


# In[12]:


#prog10
# Get a number from the user
num = float(input("Enter a number: "))

# Check if the number is positive, negative, or zero
if num > 0:
    print("The number is positive")
elif num < 0:
    print("The number is negative")
else:
    print("The number is zero")


# In[13]:


#prog12
# Get a year from the user
year = int(input("Enter a year: "))

# Check if it's a leap year
if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
    print(f"{year} is a leap year")
else:
    print(f"{year} is not a leap year")


# In[ ]:





# In[14]:


#prog11
import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook("your_excel_file.xlsx")  # Replace with your Excel file path

# Get the active sheet (you can specify a sheet name if needed)
sheet = workbook.active

# Get the cell name from the user
cell_name = input("Enter a cell name (e.g., A1, B2, C3): ")

# Check if the cell name exists in the sheet
if cell_name in sheet:
    cell = sheet[cell_name]
    value = cell.value
    print(f"The value in cell {cell_name} is: {value}")
else:
    print(f"Cell {cell_name} does not exist in the sheet")


# In[15]:


#prog13
# Function to check if a number is prime
def is_prime(num):
    if num <= 1:
        return False
    for i in range(2, int(num**0.5) + 1):
        if num % i == 0:
            return False
    return True

# Get the interval from the user
start = int(input("Enter the starting number of the interval: "))
end = int(input("Enter the ending number of the interval: "))

# Print prime numbers in the interval
print(f"Prime numbers between {start} and {end}:")
for num in range(start, end + 1):
    if is_prime(num):
        print(num)


# In[16]:


#prog14
# Get the number for which you want to display the multiplication table
num = int(input("Enter a number: "))

# Display the multiplication table
print(f"Multiplication table for {num}:")
for i in range(1, 11):
    result = num * i
    print(f"{num} x {i} = {result}")


# In[17]:


#prog 15
#prog15
import re

# Get a string from the user
text = input("Enter a text: ")

# Use regular expressions to match 'a', 'b', or 'c' anywhere in the text
if re.search(r'[abc]', text):
    print("The characters 'a', 'b', or 'c' are found in the text.")
else:
    print("The characters 'a', 'b', or 'c' are not found in the text.")



# In[18]:


#prog16
import re

# Get a string from the user
text = input("Enter a text: ")

# Use regular expressions to match the sub-pattern "name"
if re.search(r'name', text):
    print("The sub-pattern 'name' is found in the text.")
else:
    print("The sub-pattern 'name' is not found in the text.")


# In[19]:


#prog17
import re

# Get a string from the user
text = input("Enter a string: ")

# Use regular expressions to match any decimal digit within the string
if re.search(r'\d', text):
    print("The string contains at least one decimal digit.")
else:
    print("The string does not contain any decimal digits.")


# In[20]:


#prog18
import re

# Get a string from the user
text = input("Enter a string: ")

# Use regular expressions to match a letter, digit, or underscore within the string
if re.search(r'\w', text):
    print("The string contains at least one letter, digit, or underscore.")
else:
    print("The string does not contain any letters, digits, or underscores.")


# In[21]:


#prog19
# Read contents from a file
with open("input.txt", "r") as file:
    file_contents = file.read()
    print("Contents of the file:")
    print(file_contents)

# Write contents to a file
with open("output.txt", "w") as file:
    text_to_write = "This is some text to write to the file."
    file.write(text_to_write)
    print("Contents written to 'output.txt'")


# In[ ]:


#prog20
import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook("your_excel_file.xlsx")  # Replace with your Excel file path

# Get the active sheet (you can specify a sheet name if needed)
sheet = workbook.active

# Read and print contents from the spreadsheet
for row in sheet.iter_rows(values_only=True):
    print(row)


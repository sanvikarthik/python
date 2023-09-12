#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import Workbook 
from openpyxl.styles import Font

wb = Workbook() # Create a Workbook
sheet = wb.active
sheet.title = "Language" 
wb.create_sheet(title = "Capital") 

lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code =['KA', 'TS', 'TN']


sheet.cell(row = 1, column = 1).value = "State"
sheet.cell(row = 1, column = 2).value = "Language"
sheet.cell(row = 1, column = 3).value = "Code"


ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft
        
for i in range(2,5):
    sheet.cell(row = i, column = 1).value = state[i-2]
    sheet.cell(row = i, column = 2).value = lang[i-2]
    sheet.cell(row = i, column = 3).value = code[i-2]

wb.save("C:\\Users\\deepa\\OneDrive\\Desktop\\demo.xlsx")


srchCode = input("Enter state code for finding language ")
for i in range(2,5):
    data = sheet.cell(row = i, column = 3).value
    if data == srchCode:
        print("Corresponding language for code", srchCode, "is", sheet.cell(row = i, column = 2).value)

wb.close()


# In[2]:


#prog10a
from PyPDF2 import PdfWriter, PdfReader 
num1 = int(input("Enter page number from file1 ")) 
num2 = int(input("Enter page number from file2 ")) 
pdf1 = open("D:\\Even Sem 2023\\ex1.pdf ", 'rb') 
pdf2 = open("D:\\Even Sem 2023\\ex1.pdf ", 'rb') 
pdf_writer = PdfWriter() 
pdf1_reader = PdfReader(pdf1) 
page = pdf1_reader.pages[num1 - 1] 
pdf_writer.add_page(page) 
pdf2_reader = PdfReader(pdf2) 
page = pdf2_reader.pages[num2 - 1] 
pdf_writer.add_page(page) 
with open('D:\\ Even Sem 2023\\output.pdf', 'wb') as output: 
    pdf_writer.write(output) 
print("Combined successfully")


# In[ ]:


#prog10b
import json

with open("C:\\Desktop\\weather.json.txt") as f:
    data = json.load(f)

current_temp = data['main']['temp']
humidity = data['main']['humidity']
weather_desc = data['weather'][0]['description']

print(f"Current temperature: {current_temp}°C")
print(f"Humidity: {humidity}%")
print(f"Weather description: {weather_desc}")

